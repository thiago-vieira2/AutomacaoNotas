from openpyxl import Workbook, load_workbook
from datetime import date
#Leitura do arquivo 
arquivo = r"C:\Users\Thiag\OneDrive\Área de Trabalho\AutomacaoNotas\sla.xlsx" # Substituir pela do servidor 
wb = load_workbook(filename=arquivo)
sheet = wb["Planilha1"]
hoje = date.today()
hoje_formatado = hoje.strftime("%d/%m/%Y")


def AdicionarValores(Nf, Pedido, Fornecedor, Valor):
    print(sheet.max_row)
    sheet.append([hoje_formatado, Nf, Pedido, Fornecedor, Valor])
    wb.save(arquivo)
    for linha in sheet.iter_rows(values_only=True):
      print(linha)
